'''
Author: Wei Ge (University of Melbourne)
Email:  gewg046@outlook.com

The generator to create student's excel automatically
'''

from email import message
import tkinter as tk
from tkinter import *
from tkinter import messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import tkinter.filedialog as fd

student_df = pd.DataFrame()
temp_path = ""
target_path = ""

'''
Load the student list,
Return the dataframe with standard format
'''
def student_info_reader(file_path):
    global student_df

    # get the student list
    student_list = pd.read_excel(file_path)

    # # remove the useless information
    # student_list.drop(student_list.head(2).index,inplace=True) 

    # # get the column name
    # col_name = {}
    # n = 0
    # for col in student_list.loc[2]:
    #     col_name[student_list.columns.to_list()[n]] = col
    #     n += 1
    # # change the column name
    # student_list = student_list.rename(columns=col_name)
    # student_list.drop(student_list.head(1).index,inplace=True) 

    student_df = student_list
    # drop the row with student name as NaN
    student_df.dropna(subset=["中文姓名"], inplace=True)


def change_cell_style(target1, row_num, n):
    # the cell stype
    side = Side("thin")
    # add the cell style
    for i in range(1, n+1):
        target1.cell(row=row_num,column=i).border = Border(left=side, right=side, top=side, bottom=side)


def generate_single_grade_report(template_file_path, result_file_name):
    '''
    Generate the single grade report
    @param template_file_path: file path of template
    @param result_file_name: result file name
    '''
    global student_df

    '''get the template file'''
    workbook = load_workbook(template_file_path)
    template = workbook["单次Grade Report"]
    target1 = workbook.copy_worksheet(template)

    '''generate the grade report'''
    # the total number of student
    student_num = 1
    # the paging count, page out when hits 30 students
    paging_count = 1
    # the column number
    row_num = 9
    # page number
    page_number = 1
    # cell style
    side = Side("thin")
    # search all students then insert their information
    for index, row in student_df.iterrows():
        # page out
        if (paging_count == 31):
            # insert the signature
            target1.cell(row=row_num,column=5,value="Instructor Signature: ")
            target1.cell(row=row_num,column=6).border = Border(bottom=side)
            target1.cell(row=row_num,column=7).border = Border(bottom=side)
            row_num += 1
            # insert the page no.
            target1.cell(row=row_num,column=2,value="Page " + str(page_number))
            # insert date
            target1.cell(row=row_num,column=5,value="Date: ")
            target1.cell(row=row_num,column=6).border = Border(bottom=side)
            target1.cell(row=row_num,column=7).border = Border(bottom=side)
            row_num += 1

            page_number += 1
            paging_count = 1

        # get the first and last name
        firstname = row["名"]
        lastname = row["姓"]
        # get the chinese name
        chinese_name = row["中文姓名"]
        # get the birthday
        birthday = row["出生日期"]

        # insert the informations
        target1.cell(row=row_num,column=1,value=student_num)
        target1.cell(row=row_num,column=2,value=chinese_name)
        target1.cell(row=row_num,column=3,value=firstname)
        target1.cell(row=row_num,column=4,value=lastname)
        target1.cell(row=row_num,column=5,value=birthday)
        # insert the cell style
        change_cell_style(target1, row_num, 7)

        # plus the loop
        row_num += 1
        student_num += 1
        paging_count += 1
    
    # fill in the singature and date of the last page
    while (paging_count < 31):
        # insert the cell style
        change_cell_style(target1, row_num, 7)

        row_num += 1
        paging_count += 1

        if (paging_count == 31):
            # insert the signature
            target1.cell(row=row_num,column=5,value="Instructor Signature: ")
            target1.cell(row=row_num,column=6).border = Border(bottom=Side("thin"))
            target1.cell(row=row_num,column=7).border = Border(bottom=Side("thin"))
            row_num += 1
            # insert the page no.
            target1.cell(row=row_num,column=2,value="Page " + str(page_number))
            # insert date
            target1.cell(row=row_num,column=5,value="Date: ")
            target1.cell(row=row_num,column=6).border = Border(bottom=Side("thin"))
            target1.cell(row=row_num,column=7).border = Border(bottom=Side("thin"))
            row_num += 1

    # save the file
    workbook.save(filename = result_file_name)



def generate_course_grade_report(template_file_path, result_file_name):
    '''
    Generate the course grade report
    @param template_file_path: file path of template
    @param result_file_name: result file name
    '''
    global student_df

    '''get the template file'''
    workbook = load_workbook(template_file_path)
    template = workbook["Course Grade Report"]
    target1 = workbook.copy_worksheet(template)

    '''generate the grade report'''
    # the total number of student
    student_num = 1
    # the paging count, page out when hits 30 students
    paging_count = 1
    # the column number
    row_num = 8
    # page number
    page_number = 1
    # cell style
    side = Side("thin")
    # search all students then insert their information
    for index, row in student_df.iterrows():

        # page out
        if (paging_count == 16):
            row_num += 1
            # insert the page no.
            target1.cell(row=row_num,column=3,value="Page " + str(page_number))
            row_num += 1
            page_number += 1
            paging_count = 1
        student_no = row["订单号"]
        student_name = row["中文姓名"]
        student_firstname = row["名"]
        student_lastname = row["姓"]
        dob = row["出生日期"]
        # insert the informations
        target1.cell(row=row_num,column=1,value=student_num)
        target1.cell(row=row_num,column=2,value=student_no)
        target1.cell(row=row_num,column=3,value=student_name)
        target1.cell(row=row_num,column=4,value=student_firstname)
        target1.cell(row=row_num,column=5,value=student_lastname)
        target1.cell(row=row_num,column=6,value=dob)
        # insert the cell style
        change_cell_style(target1, row_num, 11)

        # plus the loop
        row_num += 1
        student_num += 1
        paging_count += 1

    # fill in the singature and date of the last page
    while (paging_count < 16):
        # insert the cell style
        change_cell_style(target1, row_num, 11)

        row_num += 1
        paging_count += 1

        if (paging_count == 16):
            row_num += 1
            # insert the page no.
            target1.cell(row=row_num,column=3,value="Page " + str(page_number))
            row_num += 1

    # save the file
    workbook.save(filename = result_file_name)


def generate_attendance(template_file_path, result_file_name):
    '''
    Generate the attendence
    @param template_file_path: file path of template
    @param result_file_name: result file name
    '''
    global student_df

    '''get the template file'''
    workbook = load_workbook(template_file_path)
    template = workbook["sheet1"]
    target1 = workbook.copy_worksheet(template)

    '''generate the grade report'''
    # the total number of student
    student_num = 1
    # the paging count, page out when hits 30 students
    paging_count = 1
    # the column number
    row_num = 5
    # page number
    page_number = 1
    # cell style
    side = Side("thin")
    # search all students then insert their information
    for index, row in student_df.iterrows():
        student_cn_name = row["中文姓名"]
        student_firstname = row["名"]
        student_lastname = row["姓"]
        student_eng_name = student_firstname + " " + student_lastname
        # insert the informations
        target1.cell(row=row_num,column=1,value=student_num)
        target1.cell(row=row_num,column=2,value=student_cn_name)
        target1.cell(row=row_num,column=3,value=student_eng_name)
        # insert the cell style
        change_cell_style(target1, row_num, 23)
        # plus the loop
        row_num += 1
        student_num += 1

    # save the file
    workbook.save(filename = result_file_name)


def gender_transfer(cn_gender):
    '''
    Transfer gender from chinese to english
    '''
    if (cn_gender == "女"):
        return "Female"
    else:
        return "Male"


def generate_roster(template_file_path, result_file_name):
    '''
    Generate the attendence
    @param template_file_path: file path of template
    @param result_file_name: result file name
    '''
    global student_df

    '''get the template file'''
    workbook = load_workbook(template_file_path)
    template = workbook["花名册"]
    target1 = workbook.copy_worksheet(template)

    '''generate the grade report'''
    # the total number of student
    student_num = 1
    # the column number
    row_num = 5
    # cell style
    side = Side("thin")
    # search all students then insert their information
    for index, row in student_df.iterrows():
        student_firstname = row["名"]
        student_lastname = row["姓"]
        student_eng_name = student_firstname + " " + student_lastname
        student_gender = gender_transfer(row["性别"])
        student_dob = row["出生日期"]
        student_home_institution = row["学校"]
        student_time_of_enrol = row["入学时间"]
        student_major = row["专业"]
        student_gpa = row["GPA"]
        student_email = row["邮箱"]
        # insert the informations
        target1.cell(row=row_num,column=1,value=student_num)
        target1.cell(row=row_num,column=2,value=student_eng_name)
        target1.cell(row=row_num,column=3,value=student_gender)
        target1.cell(row=row_num,column=4,value=student_dob)
        target1.cell(row=row_num,column=5,value=student_home_institution)
        target1.cell(row=row_num,column=6,value=student_time_of_enrol)
        target1.cell(row=row_num,column=7,value=student_major)
        target1.cell(row=row_num,column=8,value=student_gpa)
        target1.cell(row=row_num,column=9,value=student_email)
        # insert the cell style
        change_cell_style(target1, row_num, 9)
        # plus the loop
        row_num += 1
        student_num += 1

    # save the file
    workbook.save(filename = result_file_name)


def info_confirm(template_file_path, result_file_name):
    '''
    Generate the attendence
    @param template_file_path: file path of template
    @param result_file_name: result file name
    '''
    global student_df
    '''get the template file'''
    workbook = load_workbook(template_file_path)
    template = workbook["学生签字确认表"]
    target1 = workbook.copy_worksheet(template)

    '''generate the grade report'''
    # the total number of student
    student_num = 1
    # the column number
    row_num = 5
    # cell style
    side = Side("thin")
    # search all students then insert their information
    for index, row in student_df.iterrows():
        student_cn_name = row["中文姓名"]
        student_firstname = row["名"]
        student_lastname = row["姓"]
        student_gender = gender_transfer(row["性别"])
        student_dob = row["出生日期"]

        # insert the informations
        target1.cell(row=row_num,column=1,value=student_num)
        target1.cell(row=row_num,column=2,value=student_cn_name)
        target1.cell(row=row_num,column=3,value=student_firstname)
        target1.cell(row=row_num,column=4,value=student_lastname)
        target1.cell(row=row_num,column=5,value=student_gender)
        target1.cell(row=row_num,column=6,value=student_dob)
        # insert the cell style
        change_cell_style(target1, row_num, 7)
        # plus the loop
        row_num += 1
        student_num += 1

    # save the file
    workbook.save(filename = result_file_name)



def generate_all():
    '''
    The main function to generate all excels
    '''
    # read file info
    # student_info_reader("BU263 风险管理和衍生品-20220619151404.xls")
    # # single grade report
    # generate_single_grade_report("2022 July_SJTU Single Grade Report.xlsx", "Test_Single_Report.xlsx")
    # # course grade report
    # generate_course_grade_report("2022 July_SJTU Course Grade Report.xlsx", "Test_Course_Report.xlsx")
    # # attendence
    # generate_attendance("Attendence_Temp.xlsx", "Test_Attendence.xlsx")
    # # roster
    # generate_roster("Roster_Temp.xlsx", "Test_Roster.xlsx")
    # info_confirm()
 

def loadInfo(show_entry):
    '''
    Bind the information file
    '''
    file_path = fd.askopenfilename()
    show_entry.delete(0,END)
    show_entry.insert(0,file_path)
    # load information
    student_info_reader(file_path)


def loadTemp(show_entry):
    '''
    Bind the template
    '''
    global temp_path

    file_path = fd.askopenfilename()
    show_entry.delete(0,END)
    show_entry.insert(0,file_path)
    # load information
    temp_path = file_path


def bindTarget(show_entry):
    '''
    Bind the target file
    '''
    global target_path

    file_path = fd.askdirectory()
    show_entry.delete(0,END)
    show_entry.insert(0,file_path)
    # load information
    target_path = file_path


if __name__ == '__main__':
    root=Tk()
    root.title("Student List Generator")

    en1=Entry(root)
    en1.grid(row=0, column=1, columnspan=3)
    la1=Label(root, text="已选择的学生信息文件: ")
    la1.grid(row=0, column=0)   
    but1=Button(root, text="选择学生信息文件", command=lambda : loadInfo(en1))
    but1.grid(row=0, column=4)
    
    la2=Label(root,text="已选择的Template文件: ")
    la2.grid(row=1, column=0)
    en2=Entry(root)
    en2.grid(row=1,column=1,columnspan=3)
    but2=Button(root, text="选择Template文件", command=lambda : loadTemp(en2))
    but2.grid(row=1, column=4)

    la3=Label(root,text="目标文件路径：")
    la3.grid(row=2, column=0)
    en3=Entry(root)
    en3.grid(row=2,column=1,columnspan=3)
    but2=Button(root, text="选择目标文件路径", command=lambda : bindTarget(en3))
    but2.grid(row=2, column=4)
    

    but6=tk.Button(root, text=' 生成 "花名册" ', command=lambda : generate_roster(temp_path, target_path + "/自动生成_花名册.xlsx"))
    but6.grid(row=3,column=0)
    but3=Button(root,text=' 生成 "考勤表" ', command=lambda : generate_attendance(temp_path, target_path + "/自动生成_考勤表.xlsx"))
    but3.grid(row=3,column=1)
    but4=Button(root,text=' 生成 "Course Grade Report" ', command=lambda : generate_course_grade_report(temp_path, target_path + "/自动生成_Course_Grade_Report.xlsx"))
    but4.grid(row=3,column=2)
    but5=Button(root,text=' 生成 "单次Grade Report" ', command=lambda : generate_single_grade_report(temp_path, target_path + "/自动生成_单次Grade_Report.xlsx"))
    but5.grid(row=3,column=3)
    but7=tk.Button(root, text=' 生成 "学生信息确认表" ', command=lambda : info_confirm(temp_path, target_path + "/自动生成_学生信息确认表.xlsx"))
    but7.grid(row=3,column=4)


    root.mainloop()